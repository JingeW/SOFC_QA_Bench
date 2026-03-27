"""Export all question types to Excel for human evaluation."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from sofc_bench.utils.io import load_jsonl, load_yaml

# Strip "Answer:" (case-insensitive) and any following whitespace for human-facing display
_ANSWER_PREFIX_RE = re.compile(r"^\s*Answer:\s*", re.IGNORECASE)


def _strip_answer_prefix(text: str) -> str:
    """Remove leading 'Answer:' (case-insensitive) and any following whitespace."""
    if not text or not isinstance(text, str):
        return text
    return _ANSWER_PREFIX_RE.sub("", text.strip(), count=1).strip()


def _format_sc_mc_question(question_stem: str, options: dict[str, str] | None) -> str:
    """Format question and options exactly as: Question:\\n<stem>\\n\\nOptions:\\nA. ...\\nB. ...\\nC. ...\\nD. ..."""
    out = "Question:\n" + (question_stem or "").strip()
    out += "\n\nOptions:\n"
    if options:
        for letter in ("A", "B", "C", "D"):
            if letter in options:
                out += f"{letter}. {options[letter]}\n"
    return out.strip()


def _load_questions_by_type(jsonl_dir: Path) -> dict[str, dict[str, dict[str, Any]]]:
    """Load dataset JSONL per type; return {type: {question_id: row}}."""
    result: dict[str, dict[str, dict[str, Any]]] = {}
    for qtype, filename in [
        ("tf", "tf.jsonl"),
        ("single_choice", "single_choice.jsonl"),
        ("multi_choice", "multi_choice.jsonl"),
        ("open_ended", "open_ended.jsonl"),
    ]:
        path = jsonl_dir / filename
        if path.exists():
            result[qtype] = {str(r["question_id"]): r for r in load_jsonl(path)}
        else:
            result[qtype] = {}
    return result


def export_review_to_excel(
    run_dir: Path,
    output_path: Path | None = None,
    project_root: Path | None = None,
) -> Path:
    """Write all question types to one Excel workbook for human review.

    Sheets: tf, single_choice, multi_choice, open_ended.
    - Closed (tf, single_choice, multi_choice): ID, Question, Answer, GT, is_valid, is_correct.
    - SC/MC Question cell format: Question:\\n<stem>\\n\\nOptions:\\nA. ...\\nB. ...\\nC. ...\\nD. ...
    - Open_ended: ID, Question, Answer, Grade, Comments (Grade/Comments empty for expert).
    Requires parsed_outputs_<type>.jsonl and run_config_snapshot.yaml in run_dir.
    """
    run_dir = Path(run_dir)
    snapshot_path = run_dir / "run_config_snapshot.yaml"
    if not snapshot_path.exists():
        raise FileNotFoundError(f"run_config_snapshot.yaml not found: {snapshot_path}")

    snapshot = load_yaml(snapshot_path)
    dataset_id = snapshot.get("dataset_id", "sofc_v1")
    root = Path(project_root) if project_root else run_dir.parent.parent
    jsonl_dir = root / "data" / "qa_sets" / dataset_id / "jsonl"
    if not jsonl_dir.is_dir():
        raise FileNotFoundError(f"Dataset jsonl dir not found: {jsonl_dir}")

    questions_by_type = _load_questions_by_type(jsonl_dir)

    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. pip install openpyxl") from None

    wb = openpyxl.Workbook()
    closed_headers = ["ID", "Question", "Answer", "GT", "is_valid", "is_correct"]
    open_headers = ["ID", "Question", "Answer", "Grade", "Comments"]
    type_list = [("tf", closed_headers), ("single_choice", closed_headers), ("multi_choice", closed_headers), ("open_ended", open_headers)]

    for idx, (qtype, headers) in enumerate(type_list):
        parsed_path = run_dir / f"parsed_outputs_{qtype}.jsonl"
        if not parsed_path.exists():
            continue
        parsed = load_jsonl(parsed_path)
        q_by_id = questions_by_type.get(qtype, {})

        if qtype == "tf":
            rows = []
            for r in sorted(parsed, key=lambda x: (str(x.get("question_id", "")), x.get("model_id", ""))):
                qid = str(r.get("question_id", ""))
                q_row = q_by_id.get(qid, {})
                question_text = q_row.get("question") or q_row.get("text") or ""
                rows.append({
                    "Question": question_text,
                    "Answer": _strip_answer_prefix(r.get("raw_text") or r.get("parsed_answer") or ""),
                    "GT": r.get("expected") or "",
                    "is_valid": r.get("is_valid"),
                    "is_correct": r.get("is_correct"),
                })
        elif qtype in ("single_choice", "multi_choice"):
            rows = []
            for r in sorted(parsed, key=lambda x: (str(x.get("question_id", "")), x.get("model_id", ""))):
                qid = str(r.get("question_id", ""))
                q_row = q_by_id.get(qid, {})
                stem = q_row.get("question") or q_row.get("text") or ""
                opts = q_row.get("options") if isinstance(q_row.get("options"), dict) else None
                rows.append({
                    "Question": _format_sc_mc_question(stem, opts),
                    "Answer": _strip_answer_prefix(r.get("raw_text") or r.get("parsed_answer") or ""),
                    "GT": r.get("expected") or "",
                    "is_valid": r.get("is_valid"),
                    "is_correct": r.get("is_correct"),
                })
        else:
            # open_ended
            rows = []
            for r in sorted(parsed, key=lambda x: (str(x.get("question_id", "")), x.get("model_id", ""))):
                qid = str(r.get("question_id", ""))
                q_row = q_by_id.get(qid, {})
                question_text = q_row.get("question") or q_row.get("text") or ""
                answer_text = _strip_answer_prefix(r.get("parsed_answer") or r.get("raw_text") or "")
                rows.append({
                    "Question": question_text,
                    "Answer": answer_text,
                    "Grade": "",
                    "Comments": "",
                })

        if idx == 0:
            ws = wb.active
            ws.title = qtype
        else:
            ws = wb.create_sheet(title=qtype, index=idx)
        ws.append(headers)
        for i, row in enumerate(rows, start=1):
            if headers == closed_headers:
                ws.append([i, row["Question"], row["Answer"], row["GT"], row["is_valid"], row["is_correct"]])
            else:
                ws.append([i, row["Question"], row["Answer"], row["Grade"], row["Comments"]])
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 60
        if headers == closed_headers:
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 10

    out = Path(output_path) if output_path else run_dir / "human_review.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def export_open_ended_to_excel(
    run_dir: Path,
    output_path: Path | None = None,
    project_root: Path | None = None,
) -> Path:
    """Export all question types to Excel for human review (single workbook, 4 sheets)."""
    return export_review_to_excel(run_dir, output_path=output_path, project_root=project_root)
